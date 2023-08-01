import cv2
import os
import numpy as np
import openpyxl
import numpy
from matplotlib import pyplot as plt
from tqdm import tqdm
from ellipse import ellipse
from otsu import yuv_otsu
from ripness_data import tomato_ripness
from excel_clear import excel_clear
from PIL import Image
from openpyxl import load_workbook
import math
import sys   
import signal
import argparse
import warnings
import gdal
import rasterio
warnings.filterwarnings("ignore")


# https://docs.python.org/zh-tw/3/howto/argparse.html
parser = argparse.ArgumentParser()
parser.add_argument('--datasets', type=str, default='19th2_tomato', help='番茄資料集名稱')
parser.add_argument('--number', type=str, default='6th', help='第幾次拍攝')
parser.add_argument('--lane', type=str, default='3l_test', help='第幾排番茄')
args = parser.parse_args()


# path='F:/dataset/19th_tomato/6th/sequoia/8l_test'
# rgb_filepath=path+'/RGB'
# path='F:\warpping_try'
# path='F:\dataset/lin_tomato/4th/sequoia/2l_train'
path=os.path.join('F:\dataset',args.datasets,args.number,'sequoia',args.lane)
# rgb_filepath=path+'/images/RGB'
rgb_filepath=path+'/images/RGB'
ndvi_filepath=path+'/NDVI'
ndre_filepath=path+'/NDRE'
gndvi_filepath=path+'/GNDVI'
grri_filepath=path+'/GRRI'
nir_filepath=path+'/NIR'

k=path.split('/')
# rgb_filepath='F:\dataset/19th_tomato/6th/sequoia/8l_test/RGB'

labels_filepath=path+'/labels_with_ids'
#labels_filepath=rgb_filepath+'/labels_with_ids'
# excel_path='F:\warpping_try/try'+'.xls' 
# excel_path='F:\warpping_try/try'+'.xls'
# excel_path='F:\dataset/19th_tomato/svm_test/'+k[-1]+'_'+k[-3]+'.xls' 
if(args.lane[-5:]=='train'):
    excel_path=os.path.join('F:\dataset' , args.datasets , 'svm_train' , args.lane+'_'+args.number+'.xls')
elif(args.lane[-4:]=='test'):
    excel_path=os.path.join('F:\dataset' , args.datasets , 'svm_test' , args.lane+'_'+args.number+'.xls')
#F:\dataset/19th_tomato/7th\sequoia/8l_test/7th_8l_test.xlsx

# dataset=[k[-4],k[-3],k[-1]]
dataset=[args.datasets,args.number,args.lane]
#'19th_tomato','4th','3l_train'
ripness_data=tomato_ripness(dataset)
rgb_files=os.listdir(rgb_filepath)
ndvi_files=os.listdir(ndvi_filepath)
ndre_files=os.listdir(ndre_filepath)
gndvi_files=os.listdir(gndvi_filepath)
grri_files=os.listdir(grri_filepath)
nir_files=os.listdir(nir_filepath)
# print(len(rgb_files))
print(path)
pbar = tqdm(total=len(rgb_files),desc='計算各番茄的像素平均值')

# if not os.path.isdir(excel_path):
#     # 利用 Workbook 建立一個新的工作簿
#     workbook = openpyxl.Workbook()
# else:
#     workbook = openpyxl.load_workbook(excel_path)

# isExist = os.path.exists(excel_path)
# if not isExist:
#     os.makedirs(excel_path)
# 利用 Workbook 建立一個新的工作簿
workbook = openpyxl.Workbook()
# workbook = openpyxl.load_workbook(excel_path)
# 取得第一個工作表
sheet = workbook.worksheets[0]
# 設定 excel 工作表
excel_sheet=['id','red','green','blue','r-g',
            'hue','saturation','brightness','s-h',
            'l','a','b','l-a',
            'ndvi','ndre','gndvi','grri','ripness']
sheet['A1'] = excel_sheet[0]
sheet['B1'] = excel_sheet[1]
sheet['C1'] = excel_sheet[2]
sheet['D1'] = excel_sheet[3]
sheet['E1'] = excel_sheet[4]
sheet['F1'] = excel_sheet[5]
sheet['G1'] = excel_sheet[6]
sheet['H1'] = excel_sheet[7]
sheet['I1'] = excel_sheet[8]
sheet['J1'] = excel_sheet[9]
sheet['K1'] = excel_sheet[10]
sheet['L1'] = excel_sheet[11]
sheet['M1'] = excel_sheet[12]
sheet['N1'] = excel_sheet[13]
sheet['O1'] = excel_sheet[14]
sheet['P1'] = excel_sheet[15]
sheet['Q1'] = excel_sheet[16]
sheet['R1'] = excel_sheet[17]
column=2
# Correction_coefficient=0.7 #校正係數


# 儲存檔案
workbook.save(excel_path)
# def write_excel():
tt=0
filtration_capacity=95
def filter(image,tt):
    img_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # 先轉成灰階處理
    maxium=np.percentile(img_gray, filtration_capacity, interpolation='midpoint') #https://www.796t.com/content/1550429124.html
    minimum=np.percentile(img_gray, (100-filtration_capacity)+100*(1-0.25*math.pi), interpolation='midpoint')
    #https://blog.csdn.net/JNingWei/article/details/77747959
    ret, thresh1 = cv2.threshold(img_gray, maxium, 255, 4)  # 利用 threshold 過濾出 #type = 4 ：此时小于阈值的像素点保持原色，大于阈值的像素点置0
    ret, thresh2 = cv2.threshold(img_gray, minimum, 255, 3)  # type = 3 ：此时大于阈值的像素点置填充色，小于阈值的像素点置0
    # concat_pic = np.concatenate([img_gray, thresh], axis=1)
    img = cv2.cvtColor(thresh1, cv2.COLOR_GRAY2BGR)  
    res = cv2.bitwise_and(image, image,mask=thresh1)
    res = cv2.bitwise_and(res, res,mask=thresh2)
    # if(tt<=300):
    #     cv2.imwrite('./crop_images/6th_85/'+str(tt)+'.jpg', res)
    
    return res





#切照片
def crop_images(x,y,w,h,img):
    
    crop_img = img[y-int(0.5*h):y+int(0.5*h), x-int(0.5*w):x+int(0.5*w)]
    
    return crop_img
#切照片
def crop_tif_images(x,y,w,h,img):
    
    
    crop_img = img.crop((x-int(0.5*w),y-int(0.5*h),x+int(0.5*w),y+int(0.5*h)))
    return crop_img


for g in range (len(rgb_files)):
    # print(rgb_files[g][-4:])
    if(rgb_files[g][-3:]=='JPG'):
        # try:
            if os.path.isfile(labels_filepath+'/'+rgb_files[g][:-4]+'.txt')==False:
                f= open(labels_filepath+'/'+rgb_files[g][:-4]+'.txt',"w+")
                f.close()
            f=open(labels_filepath+'/'+rgb_files[g][:-4]+'.txt','r')
            rgb_image=cv2.imread(rgb_filepath+'/'+rgb_files[g])
            rgb_image=cv2.resize(rgb_image, (1280,960))
            # original_image=Image.open(rgb_filepath+'/'+rgb_files[g])
            # ndvi_image=Image.open(ndvi_filepath+'/'+ndvi_files[g])
            # ndre_image=Image.open(ndre_filepath+'/'+ndre_files[g])
            # gndvi_image=Image.open(gndvi_filepath+'/'+gndvi_files[g])
            # grri_image=Image.open(grri_filepath+'/'+grri_files[g])
            # nir_image=Image.open(nir_filepath+'/'+nir_files[g])
            
            ndvi_image=cv2.imread(ndvi_filepath+'/'+ndvi_files[g],2)
            ndre_image=cv2.imread(ndre_filepath+'/'+ndre_files[g],2)
            gndvi_image=cv2.imread(gndvi_filepath+'/'+gndvi_files[g],2)
            grri_image=cv2.imread(grri_filepath+'/'+grri_files[g],2)
            # print(ndvi_image.size)
            # with rasterio.open(ndvi_filepath+'/'+ndvi_files[g]) as src:
            #     ndvi_image = src.read()
            # ndvi_image = np.dot(ndvi_image[:3], [0.2989, 0.5870, 0.1140])

            # with rasterio.open(ndre_filepath+'/'+ndre_files[g]) as src:
            #     ndre_image = src.read()
            # ndre_image = np.dot(gndvi_image[:3], [0.2989, 0.5870, 0.1140])

            # with rasterio.open(gndvi_filepath+'/'+gndvi_files[g]) as src:
            #     gndvi_image = src.read()
            # gndvi_image = np.dot(ndvi_image[:3], [0.2989, 0.5870, 0.1140])

            # with rasterio.open(grri_filepath+'/'+grri_files[g]) as src:
            #     grri_image = src.read()
            # grri_image = np.dot(grri_image[:3], [0.2989, 0.5870, 0.1140])
            # print(ndvi_image)
            
            # nir_image=Image.open(nir_filepath+'/'+nir_files[g])
            # print(ndvi_filepath+'/'+ndvi_files[g]) 
              
            

            # original_size = original_image.size
            for line in f.readlines():
                s = line.split(' ')
                # print(nir_filepath+'/'+nir_files[g])
                Correction_coefficient=0.9
                
                # id=int(s[1])                
                # x_center=float(s[2])*(ndvi_image.size[0])
                # y_center=float(s[3])*(ndvi_image.size[1])
                # width=float(s[4])*(ndvi_image.size[0])*Correction_coefficient
                # height=float(s[5])*(ndvi_image.size[1])*Correction_coefficient

                id=int(s[1])                
                x_center=float(s[2])*(ndvi_image.shape[1])
                y_center=float(s[3])*(ndvi_image.shape[0])
                width=float(s[4])*(ndvi_image.shape[1])*Correction_coefficient
                height=float(s[5])*(ndvi_image.shape[0])*Correction_coefficient

                
                # image2=ellipse(image2)#切成橢圓
                # print(image2.shape)
                # print(x_center-0.5*width,y_center-0.5*height,x_center+0.5*width,y_center+0.5*height,ndvi_image.size[0],ndvi_image.size[1])
                

                # sndvi=ndre=0
                
                if(((x_center-0.5*width)>ndvi_image.shape[1]*0.1)&((y_center-0.5*height)>ndvi_image.shape[0]*0.05)
                    &((x_center+0.5*width)<ndvi_image.shape[1]*0.9)&((y_center+0.5*height)<ndvi_image.shape[0]*0.95)):

                    image2=crop_images(int(x_center),int(y_center),int(width),int(height),ndvi_image)
                    image3=crop_images(int(x_center),int(y_center),int(width),int(height),ndre_image)
                    image4=crop_images(int(x_center),int(y_center),int(width),int(height),gndvi_image)
                    image5=crop_images(int(x_center),int(y_center),int(width),int(height),grri_image)
                    # nir_image=crop_images(int(x_center),int(y_center),int(width),int(height),nir_image)
                    


                    # Correction_coefficient=0.7 #校正係數
                    
                    id=int(s[1])
                    x_center=float(s[2])*(rgb_image.shape[1])
                    y_center=float(s[3])*(rgb_image.shape[0])
                    width=float(s[4])*(rgb_image.shape[1])*Correction_coefficient
                    height=float(s[5])*(rgb_image.shape[0])*Correction_coefficient

                    image=crop_images(int(x_center),int(y_center),int(width),int(height),rgb_image)
                    image=ellipse(image)#切成橢圓
                    # image=filter(image,tt)
                    image=yuv_otsu(image)

                    
                    tt=tt + 1
                    # print(id,x_center-0.5*width,y_center-0.5*height,x_center+0.5*width,y_center+0.5*height,original_image.shape[1],original_image.shape[0])
                    # size = image.shape
                    h, w = image.shape[:2]
                    # print(image.size)
                    # image=cv2.imread(rgb_filepath+'/'+rgb_files[g])
                    HSV=cv2.cvtColor(image,cv2.COLOR_BGR2HSV)
                    RGB=cv2.cvtColor(image,cv2.COLOR_BGR2RGB)
                    LAB=cv2.cvtColor(image,cv2.COLOR_BGR2LAB)
                    
                    
                    # cv2.imwrite('F:/dataset/19th_tomato/6th\sequoia/3l_train/crop/'+rgb_files[g][:-4]+'_'+s[1]+'.jpg',image)
                    # print('Size:'+str(size))
                    # print(size)
                    # exist = (image[0] != 0)
                    # area = exist.sum()
                    
                    area=0
                    red=green=blue=0
                    
                    hue=0
                    saturation=0
                    brightness=0
                    l=0
                    a=0
                    b=0
                    ndvi=0
                    ndre=0
                    gndvi=0
                    grri=0

                    
                     
                    for i in range (w):
                        for j in range(h):
                            
                                
                            if(RGB[j,i][0]!=0 and RGB[j,i][1]!=0 and RGB[j,i][2]!=0):
                                area=area+1
                                red=red+RGB[j,i][0]
                                green=green+RGB[j,i][1]
                                blue=blue+RGB[j,i][2]
                                # if(RGB[j,i][0]>0 and RGB[j,i][1]>0 and RGB[j,i][2]>0):
                                
                                #OpenCV uses H: 0-179, S: 0-255, V: 0-255
                                hue=hue+HSV[j,i][0]
                                saturation=saturation+HSV[j,i][1]
                                brightness=brightness+HSV[j,i][2]

                                l=l+LAB[j,i][0]
                                a=a+LAB[j,i][1]
                                b=b+LAB[j,i][2]

                                
                    # print(w,h)
                    # print(ndvi_image.size)
                    # image1 = numpy.array(ndvi_image)
                    

                    # image1=np.array(ndvi_image)
                    # image2=np.array(ndre_image)
                    # image3=np.array(gndvi_image)
                    # image4=np.array(grri_image)
                    
                    # image1 = cv2.cvtColor(ndvi_image, cv2.COLOR_BGR2GRAY)  
                    # image2 = cv2.cvtColor(ndre_image, cv2.COLOR_BGR2GRAY)
                    # image3 = cv2.cvtColor(gndvi_image, cv2.COLOR_BGR2GRAY)
                    # image4 = cv2.cvtColor(grri_image, cv2.COLOR_BGR2GRAY) 
                    # print('\n')
                    # print(h, w) 
                    # print(ndvi_image.shape)   
                    for i in range (w):
                        for j in range(h):
                            
                                
                            if(RGB[j,i][0]!=0 and RGB[j,i][1]!=0 and RGB[j,i][2]!=0):
  
                                ndvi=ndvi+image2[j,i]
                                ndre=ndre+image3[j,i]
                                gndvi=gndvi+image4[j,i]
                                grri=grri+image5[j,i]
                    # print('hello')

                    

                    if(area!=0):
                        averge_red=(red/area)
                        averge_green=(green/area)
                        averge_blue=(blue/area)
                        rgb_output=[averge_red,averge_green,averge_blue]
                        averge_hue=(hue/area)
                        averge_saturation=(saturation/area)
                        averge_brightness=(brightness/area)
                        hsv_output=[averge_hue,averge_saturation,averge_brightness]

                        averge_l=(l/area)
                        averge_a=(a/area)
                        averge_b=(b/area)

                        average_ndvi=ndvi/area*256
                        average_ndre=ndre/area*256
                        average_gndvi=gndvi/area*256
                        average_grri=grri/area*256
                        
                        # hsv_output=[averge_hue,averge_saturation,averge_brightness]

                        # wb = openpyxl.load_workbook(excel_path, data_only=True)
                        # data = [id,averge_red,averge_green,averge_blue,averge_hue,averge_saturation,averge_brightness]   # 二維陣列資料
                        # for i in data:
                        #     s3.append(i)

                        #將資料寫入excel
                        # print('q')
                        
                        sheet['B'+str(column)] =averge_red
                        sheet['C'+str(column)] =averge_green
                        sheet['D'+str(column)] =averge_blue
                        sheet['E'+str(column)] =averge_red - averge_green
                        sheet['F'+str(column)] =averge_hue
                        sheet['G'+str(column)] =averge_saturation
                        sheet['H'+str(column)] =averge_brightness
                        sheet['I'+str(column)] =averge_saturation - averge_hue
                        sheet['J'+str(column)] =averge_l
                        sheet['K'+str(column)] =averge_a
                        sheet['L'+str(column)] =averge_b
                        sheet['M'+str(column)] =averge_l - averge_a
                        sheet['N'+str(column)] =average_ndvi
                        sheet['O'+str(column)] =average_ndre
                        sheet['P'+str(column)] =average_gndvi
                        sheet['Q'+str(column)] =average_grri
                    sheet['A'+str(column)] =id
                    
                    # sheet['M'+str(column)] =ndvi
                    # sheet['N'+str(column)] =ndre
                    # sheet['O'+str(column)] =averge_saturation
                    # sheet['P'+str(column)] =averge_brightness
                    if(math.isnan(ripness_data[id])==False):
                        sheet['R'+str(column)] =ripness_data[id]
                    # column=column+1
                    # excel_sheet=['id','red','green','blue','r-g',
                    #             'hue','saturation','brightness','s-h',
                    #             'l','a','b','l-a',
                    #               'ndvi','ndre','ripness']                               
                    column=column+1
                
                # https://python-ecw.com/2021/04/28/python-opencv%E7%95%AB%E5%9C%96/
                # img = cv2.ellipse(img, (x_center,y_center), (70,30), 15, 0, 360, (0,0,255), -1)
                
               



                        
                    
        # except:
        #     pass


    workbook.save(excel_path)
    pbar.update()

excel_clear(excel_path)

pbar.close()
