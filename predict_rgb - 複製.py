import cv2
import os
import numpy as np
import openpyxl
import numpy
from matplotlib import pyplot as plt
# import matplotlib.pyplot as plt
from tqdm import tqdm
from ellipse import ellipse
from otsu import yuv_otsu
from excel_clear import predict_excel_clear
from voting import voting
from PIL import Image
from openpyxl import load_workbook
import math
import sys   
import signal
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import pyshine as ps
import seaborn as sns
from sklearn import svm
from sklearn import datasets
from sklearn.datasets import load_iris
import joblib #https://ithelp.ithome.com.tw/articles/10197575
from sklearn.model_selection import train_test_split
from sklearn.decomposition import PCA
import warnings
warnings.filterwarnings("ignore")


# path='F:\dataset/19th_tomato/6th\sequoia\8l_test'
# predict_filepath='F:\dataset/19th_tomato\sequoia_mot/6th/tracks/img1.txt'
# k=path.split('/')
# r=predict_filepath.split('/')
# excel_path='F:\dataset/19th_tomato\sequoia_mot/6th/predict.xlsx' 
# video_path='F:\dataset/19th_tomato\sequoia_mot/6th/test.mp4'

dataset_name='19th_tomato'
data=['7th','8l_test']

# path='F:\dataset/'+str(dataset_name)+'/'+str(data[0])+'\sequoia/'+str(data[1])
# predict_filepath='F:\dataset/'+str(dataset_name)+'\sequoia_mot/'+str(data[0])+'/tracks/img1.txt'
# k=path.split('/')
# r=predict_filepath.split('/')
# excel_path='F:\dataset/'+str(dataset_name)+'\sequoia_mot/'+str(data[0])+'/predict.xlsx' 
# video_path='F:\dataset/'+str(dataset_name)+'\sequoia_mot/'+str(data[0])+'/test.mp4'

path=os.path.join('F:\dataset',str(dataset_name),str(data[0]),'sequoia',str(data[1]))
predict_filepath=os.path.join('F:\dataset',str(dataset_name),'sequoia_mot',str(data[0]),'tracks','strongsort','img1.txt')
excel_path=os.path.join('F:\dataset',str(dataset_name),'sequoia_mot',str(data[0]),'predict.xlsx')
video_path=os.path.join('F:\dataset',str(dataset_name),'sequoia_mot',str(data[0]),'test.mp4')

svm_weight = joblib.load('F:\dataset\clf.pkl')  # 選擇要使用的權重檔

# k=path.split('/')
# r=predict_filepath.split('/')
print(path)

# rgb_filepath=path+'/images/RGB'
rgb_filepath=path+'/images/RGB'
ndvi_filepath=path+'/NDVI'
ndre_filepath=path+'/NDRE'
gndvi_filepath=path+'/GNDVI'
grri_filepath=path+'/GRRI'

# k=path.split('/')

rgb_files=os.listdir(rgb_filepath)
ndvi_files=os.listdir(ndvi_filepath)
ndre_files=os.listdir(ndre_filepath)
gndvi_files=os.listdir(gndvi_filepath)
grri_files=os.listdir(grri_filepath)

#labels_filepath=rgb_filepath+'/labels_with_ids'
# excel_path='F:\warpping_try/try'+'.xls' 
# excel_path='F:\warpping_try/try'+'.xls'

# versus_component=['red','green','r-g','g-b']
# versus_component=['hue','saturation','brightness','s-h']
# versus_component=['l','a','b','l-a']
# versus_component=['r-g','g-b']
# versus_component=['r-g','green','s-h','l-a']
# versus_component=['r-g','green','s-h','l-a','ndvi']
# versus_component=['r-g','s-h','l-a']
# versus_component=['r-g','l-a']
# versus_component=['r-g','hue']
# versus_component=['r-g','g-b','s-h','l-a']
# versus_component=['r-g','hue','saturation','brightness']
# versus_component=['red','green','blue','ndvi']
# versus_component=['red','green','blue','ndre']
# versus_component=['hue','saturation','brightness','ndvi']
# versus_component=['hue','saturation','brightness','ndre']
# versus_component=['red','hue','r-g']
# versus_component=['ndvi','ndre']
# versus_component=['ndvi','gndvi','grri']
versus_component=['r-g','green','s-h','l-a','ndvi','gndvi','grri']
# versus_component=['red','green','r-g','g-b','hue','saturation','brightness','s-h','l','a','b','l-a']
#feature_columns = ['SepalLengthCm', 'SepalWidthCm', 'PetalLengthCm','PetalWidthCm']

workbook = openpyxl.Workbook()
#workbook = openpyxl.load_workbook(excel_path)
# 取得第一個工作表
sheet = workbook.worksheets[0]
# 設定 excel 工作表
excel_sheet=['frame','id','red','green','r-g','g-b',
            'hue','saturation','brightness','s-h',
            'l','a','b','l-a',
            'ndvi','ndre','gndvi','grri','ripness',
            '','left','up','width','height']
sheet['A1'] = excel_sheet[0]
sheet['B1'] = excel_sheet[1]
sheet['C1'] = excel_sheet[2]
sheet['D1'] = excel_sheet[3]
sheet['E1'] = excel_sheet[4]
sheet['F1'] = excel_sheet[5]
sheet['G1'] = excel_sheet[6]
sheet['H1'] = excel_sheet[7]
sheet['I1'] = excel_sheet[8]
sheet['J1'] = excel_sheet[9]
sheet['K1'] = excel_sheet[10]
sheet['L1'] = excel_sheet[11]
sheet['M1'] = excel_sheet[12]
sheet['N1'] = excel_sheet[13]
sheet['O1'] = excel_sheet[14]
sheet['P1'] = excel_sheet[15]
sheet['Q1'] = excel_sheet[16]
sheet['R1'] = excel_sheet[17]
sheet['S1'] = excel_sheet[18]
sheet['T1'] = excel_sheet[19]
sheet['U1'] = excel_sheet[20]
sheet['V1'] = excel_sheet[21]
sheet['W1'] = excel_sheet[22]
sheet['X1'] = excel_sheet[23]
column=2
Correction_coefficient=0.7 #校正係數


# 儲存檔案
workbook.save(excel_path)
# def write_excel():

workbook.close()
tt=0
filtration_capacity=95
def filter(image,tt):
    img_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # 先轉成灰階處理
    maxium=np.percentile(img_gray, filtration_capacity, interpolation='midpoint') #https://www.796t.com/content/1550429124.html
    minimum=np.percentile(img_gray, (100-filtration_capacity)+100*(1-0.25*math.pi), interpolation='midpoint')
    #https://blog.csdn.net/JNingWei/article/details/77747959
    ret, thresh1 = cv2.threshold(img_gray, maxium, 255, 4)  # 利用 threshold 過濾出 #type = 4 ：此时小于阈值的像素点保持原色，大于阈值的像素点置0
    ret, thresh2 = cv2.threshold(img_gray, minimum, 255, 3)  # type = 3 ：此时大于阈值的像素点置填充色，小于阈值的像素点置0
    # concat_pic = np.concatenate([img_gray, thresh], axis=1)
    img = cv2.cvtColor(thresh1, cv2.COLOR_GRAY2BGR)  
    res = cv2.bitwise_and(image, image,mask=thresh1)
    res = cv2.bitwise_and(res, res,mask=thresh2)

    # if not os.path.isdir('./crop_images/lin/5th_95/'):
    #     os.mkdir('./crop_images/lin/5th_95/')
    # if(tt<=500):
    #     cv2.imwrite('./crop_images/lin/5th_95/'+str(tt)+'.jpg', res)
    
    return res

def mask(image,tt):
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    #部分2：
    # 在HSV空间中定义葉子部分
    lower_leaf = np.array([20, 0, 0])
    upper_leaf = np.array([180, 250, 200]) #要濾掉的
    # 在HSV空间中定义绿色
    lower_reflect = np.array([0, 0, 252])
    upper_reflect = np.array([180, 255, 255])  #要濾掉的
    # 在HSV空间中定义红色
    lower_red = np.array([0, 0, 160])
    upper_red = np.array([15, 255, 245])
    
    #部分3：
    # 从HSV图像中截取出蓝色、绿色、红色，即获得相应的掩膜
    # cv2.inRange()函数是设置阈值去除背景部分，得到想要的区域
    leaf = cv2.inRange(hsv, lower_leaf, upper_leaf)
    reflect = cv2.inRange(hsv, lower_reflect, upper_reflect)
    remained1 = cv2.bitwise_not(leaf)#逐位元not邏輯運算1
    remained2 = cv2.bitwise_not(reflect)#逐位元not邏輯運算1
    #部分4：
    # 将原图像和mask(掩膜)进行按位与
    remain=cv2.bitwise_and(frame,frame,mask=remained1)
    remain=cv2.bitwise_and(remain,remain,mask=remained2)
    
    #部分5:将BGR空间下的图片转换成RGB空间下的图片
    frame = frame[:,:,::-1]
    remain = remain[:,:,::-1]
    return remain




#切照片 
def crop_images(x,y,w,h,img):
    
    crop_img = img[y-int(0.5*h):y+int(0.5*h), x-int(0.5*w):x+int(0.5*w)]
    
    return crop_img
#切照片
def crop_tif_images(x,y,w,h,img):
    
    
    crop_img = img.crop((x-int(0.5*w),y-int(0.5*h),x+int(0.5*w),y+int(0.5*h)))
    return crop_img

pbar = tqdm(total=len(rgb_files),desc='計算各番茄的像素平均值')
for g in range (len(rgb_files)):
    # print(rgb_files[g][-4:])
    if(rgb_files[g][-3:]=='JPG' or rgb_files[g][-3:]=='jpg'):
        #try:
            f=open(predict_filepath,'r')
            original_image=cv2.imread(rgb_filepath+'/'+rgb_files[g])
            original_image=cv2.resize(original_image, (256, 256), interpolation=cv2.INTER_AREA)
            ndvi_image=Image.open(ndvi_filepath+'/'+ndvi_files[g])  
            ndre_image=Image.open(ndre_filepath+'/'+ndre_files[g])
            gndvi_image=Image.open(gndvi_filepath+'/'+gndvi_files[g])
            grri_image=Image.open(grri_filepath+'/'+grri_files[g])

            # original_size = original_image.size
            for line in f.readlines():
                s = line.split(',')
                s = list(map(int, s))#string轉int
                frame=s[0]
                id=s[1]
                width=int(s[4])
                height=int(s[5])
                left=int(s[2]+(1-Correction_coefficient)/2*width)
                up=int(s[3]+(1-Correction_coefficient)/2*height)
                width=int(s[4]*Correction_coefficient)
                height=int(s[5]*Correction_coefficient)
                
                # print(id,x_center-0.5*width,y_center-0.5*height,x_center+0.5*width,y_center+0.5*height,original_image.shape[1],original_image.shape[0])



                # https://python-ecw.com/2021/04/28/python-opencv%E7%95%AB%E5%9C%96/
                # img = cv2.ellipse(img, (x_center,y_center), (70,30), 15, 0, 360, (0,0,255), -1)

                if((frame==g)&(left>=0)&(up>=0)&(0<(left+width)<original_image.shape[1])&(0<(up+height)<original_image.shape[0])):
                    # print(id,x_center-0.5*width,y_center-0.5*height,x_center+0.5*width,y_center+0.5*height,original_image.shape[1],original_image.shape[0])
                    image = original_image[up:up+height,left:left+width]
                    image=ellipse(image)#切成橢圓
                    image=filter(image,tt)
                    tt=tt + 1
                    size = image.shape
                    # print(image.size)
                    # image=cv2.imread(rgb_filepath+'/'+rgb_files[g])
                    HSV=cv2.cvtColor(image,cv2.COLOR_BGR2HSV)
                    RGB=cv2.cvtColor(image,cv2.COLOR_BGR2RGB)
                    LAB=cv2.cvtColor(image,cv2.COLOR_BGR2LAB)
                    
                    
                    # cv2.imwrite('F:/dataset/19th_tomato/6th\sequoia/3l_train/crop/'+rgb_files[g][:-4]+'_'+s[1]+'.jpg',image)
                    # area=size[0]*size[1]*0.25*math.pi - size[0]*size[1]*(100-filtration_capacity)*0.01
                    # if(area==0):
                    #     print(rgb_files[g][:-4]+' '+str(id))

                    red=0
                    green=0
                    blue=0
                    for i in range (size[1]):
                        for j in range(size[0]):
                            red=red+RGB[j,i][0]
                            green=green+RGB[j,i][1]
                            blue=blue+RGB[j,i][2]
                            if(RGB[j,i][0]!=0 and RGB[j,i][1]!=0 and RGB[j,i][2]!=0):
                                area=area+1

                    if(area==0):
                        print(rgb_files[g][:-4]+' '+str(id))

                    

                    averge_red=red/area
                    averge_green=green/area
                    averge_blue=blue/area
                    rgb_output=[averge_red,averge_green,averge_blue]
                    # print('RGB:'+str(rgb_output))    

                    #OpenCV uses H: 0-179, S: 0-255, V: 0-255
                    hue=0
                    saturation=0
                    brightness=0
                    for i in range (size[1]):
                        for j in range(size[0]):
                            hue=hue+HSV[j,i][0]/180
                            saturation=saturation+HSV[j,i][1]/256
                            brightness=brightness+HSV[j,i][2]/256



                    averge_hue=(hue/area)*180
                    averge_saturation=(saturation/area)*256
                    averge_brightness=(brightness/area)*256
                    hsv_output=[averge_hue,averge_saturation,averge_brightness]
                    # print('HSV:'+str(hsv_output))

                    l=0
                    a=0
                    b=0
                    for i in range (size[1]):
                        for j in range(size[0]):
                            l=l+LAB[j,i][0]/256
                            a=a+LAB[j,i][1]/256
                            b=b+LAB[j,i][2]/256



                    averge_l=(l/area)*256
                    averge_a=(a/area)*256
                    averge_b=(b/area)*256


                    

                    # wb = openpyxl.load_workbook(excel_path, data_only=True)
                    # data = [id,averge_red,averge_green,averge_blue,averge_hue,averge_saturation,averge_brightness]   # 二維陣列資料
                    # for i in data:
                    #     s3.append(i)

                    #將資料寫入excel
                    # print('q')
                    sheet['A'+str(column)] =frame
                    sheet['B'+str(column)] =id
                    sheet['C'+str(column)] =averge_red
                    sheet['D'+str(column)] =averge_green
                    sheet['E'+str(column)] =averge_red - averge_green
                    sheet['F'+str(column)] =averge_green - averge_blue
                    sheet['G'+str(column)] =averge_hue
                    sheet['H'+str(column)] =averge_saturation
                    sheet['I'+str(column)] =averge_brightness
                    sheet['J'+str(column)] =averge_saturation - averge_hue
                    sheet['K'+str(column)] =averge_l
                    sheet['L'+str(column)] =averge_a
                    sheet['M'+str(column)] =averge_b
                    sheet['N'+str(column)] =averge_l - averge_a
                    

                    sheet['U'+str(column)] =s[2]
                    sheet['V'+str(column)] =s[3]
                    sheet['W'+str(column)] =s[4]
                    sheet['X'+str(column)] =s[5]
                    # if(math.isnan(ripness_data[id])==False):
                    #     sheet['K'+str(column)] =ripness_data[id]
                    # column=column+1

                    

                    
                Correction_coefficient=0.5
                frame=s[0]
                id=int(s[1])                
                # width=int(s[4]*1280/4608*Correction_coefficient)
                # height=int(s[5]*960/3456*Correction_coefficient)
                # left=int(s[2]*1280/4608+(1-Correction_coefficient)*width)
                # up=int(s[3]*960/3456+(1-Correction_coefficient)*height)
                width=int(width*1280/4608)
                height=int(height*960/3456)
                left=int(left*1280/4608)
                up=int(up*960/3456)


                
                
                

                ndvi=ndre=0
                
                if((frame==g)&(left>=ndvi_image.size[0]*0.1)&(up>=ndvi_image.size[0]*0.05)&(0<(left+width)<ndvi_image.size[0]*0.9)&(0<(up+height)<ndvi_image.size[1]*0.95)):
                    # if(frame==99):
                    #     print((left,up,left+width,up+height))
                    
                    image2 = ndvi_image.crop((left,up,left+width,up+height))
                    image3 = ndre_image.crop((left,up,left+width,up+height))
                    image4 = gndvi_image.crop((left,up,left+width,up+height))
                    image5 = grri_image.crop((left,up,left+width,up+height))
                    # image2=ellipse(image2)#切成橢圓
                    average_ndvi=np.nanmean(image2)
                    average_ndre=np.nanmean(image3)
                    average_gndvi=np.nanmean(image4)
                    average_grri=np.nanmean(image5)
                    
                    sheet['O'+str(column)] =average_ndvi
                    sheet['P'+str(column)] =average_ndre
                    sheet['Q'+str(column)] =average_gndvi
                    sheet['R'+str(column)] =average_grri
                    column=column+1
                    


                        
                    
        #except:
            #pass


    workbook.save(excel_path)
    pbar.update()
predict_excel_clear(excel_path)
workbook.close()


pbar.close()

# data_xls = pd.read_excel(excel_path)
# data_xls.to_csv('F:\dataset/19th_tomato/predict.csv', encoding='utf-8')
print('預測成熟度')
workbook = openpyxl.load_workbook(excel_path)
# test_data=pd.read_csv('F:\dataset/19th_tomato/predict.csv')
test_data = pd.read_excel(excel_path)
#https://ithelp.ithome.com.tw/articles/10197575


print('versus_component='+str(versus_component))

new = test_data[versus_component].values
predict_class = test_data['ripness'].values
pca2 = PCA(n_components=len(versus_component), iterated_power=1)
# pca = PCA(n_components=2, iterated_power=1)
# test_reduced = pca2.fit_transform(new)
predicted=svm_weight.predict(new)
sheet = workbook.worksheets[0]
column=2




for i in range(len(predicted)):
    sheet['S'+str(column)] =predicted[i]
    column=column+1
workbook.save(excel_path)

voting(excel_path)

print('生成影片')
filelist=os.listdir(rgb_filepath)
size = (4608,3456)
videowrite = cv2.VideoWriter(video_path,-1,10,size)
column=0
excel_data = pd.read_excel(excel_path)
predict_frame= excel_data['frame'].values
predict_id= excel_data['id'].values
predict_ripness = excel_data['ripness'].values
left = excel_data['left'].values
up = excel_data['up'].values
width = excel_data['width'].values
height = excel_data['height'].values

m=[0,0,0,0,0] #[0,-1,-2,-3,-4]

pbar = tqdm(total=len(filelist),desc='合成影片')
id_max=0
for i in range(len(filelist)):
    img=cv2.imread(rgb_filepath+'/'+filelist[i])  
    # id_background = np.zeros((150,300,3), dtype='uint8')
    for column in range (len(predict_frame)):
        if(predict_frame[column]==i):
            #cv2.rectangle(影像, 頂點座標, 對向頂點座標, 顏色, 線條寬度)
            #https://blog.gtwang.org/programming/opencv-drawing-functions-tutorial/
            #-1到-4分別為[紅、橙、黃、綠](b,g,r)
            if(predict_ripness[column]==-1):
                # cv2.putText(img, text, org, fontFace, fontScale, color, thickness, lineType)  #cv2.putText(img, text, org, fontFace, fontScale, color, thickness, lineType)
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 0, 255), 10)
                
            elif(predict_ripness[column]==-2):
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 165, 255), 10)
                
            elif(predict_ripness[column]==-3):
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 255, 255), 10)
                
            elif(predict_ripness[column]==-4):
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 255, 0), 10)
                
        if(predict_frame[column]==i):
            if(predict_id[column]>id_max):
                id_max=predict_id[column]
                # print(predict_id[column],predict_ripness[column])
                if(predict_ripness[column]==-1):
                    
                    m[1]=m[1]+1
                elif(predict_ripness[column]==-2):
                    
                    m[2]=m[2]+1
                elif(predict_ripness[column]==-3):
                    
                    m[3]=m[3]+1
                elif(predict_ripness[column]==-4):
                    
                    m[4]=m[4]+1
            
            # print(m)

            # 文字內容

    text = 'M1:'+str(m[1])+'  M2:'+str(m[2])+'  M3:'+str(m[3])+'  M4:'+str(m[4])
        # text ='hello world'

        # 加入文字方塊
    img = ps.putBText(
            img,                            # 原始影像
            text,                             # 文字內容
            text_offset_x = 20,               # X 軸座標
            text_offset_y = 20,               # Y 軸座標
            vspace = 20,                      # 縱向空間
            hspace = 20,                      # 橫向空間
            font_scale = 6.0,                 # 字型大小
            background_RGB = (228, 225, 222), # 背景顏色
            text_RGB = (70, 90, 35)           # 文字顏色 
        )    

            
            
    videowrite.write(img)
    # m=[0,0,0,0,0]
    pbar.update()
pbar.close()
print('完成')