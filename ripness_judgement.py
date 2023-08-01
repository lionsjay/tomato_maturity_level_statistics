import openpyxl
from ripness_data import tomato_ripness
# excel_filepath = 'F:\dataset/19th_tomato/7th\sequoia/8l_test/test.xlsx'
excel_filepath = 'F:\warpping_try/try.xlsx'
wb = openpyxl.load_workbook(excel_filepath)

wb.active = 0
ws = wb.active
s=8

#ws.cell(column=2, row=3).value = 999

#row:直線
#column:橫線
id=1
ripness=11
ripness_data=tomato_ripness(10)

for i in range(2,ws.max_row+1):
    ws.cell(row=i, column=ripness).value=ripness_data[int(ws.cell(row=i, column=1).value)]
    

# for i in range(1,ws.max_row+1):
#     if(ws.cell(row=i, column=1).value == 0):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 1):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 2):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 3):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 4):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 5):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 6):
#         ws.cell(row=i, column=ripness).value = -2
#     elif(ws.cell(row=i, column=1).value == 7):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 8):
#         ws.cell(row=i, column=ripness).value = -2
#     elif(ws.cell(row=i, column=1).value == 9):
#         ws.cell(row=i, column=ripness).value = -2
#     elif(ws.cell(row=i, column=1).value == 10):
#         ws.cell(row=i, column=ripness).value = -1  

#     elif(ws.cell(row=i, column=1).value == 11):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 12):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 13):
#         ws.cell(row=i, column=ripness).value = -1
#     elif(ws.cell(row=i, column=1).value == 14):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 15):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 16):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 17):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 18):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 19):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 20):
#         ws.cell(row=i, column=ripness).value = 0

#     elif(ws.cell(row=i, column=1).value == 21):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 22):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 23):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 24):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 25):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 26):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 27):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 28):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 29):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 30):
#         ws.cell(row=i, column=ripness).value = 0 

#     elif(ws.cell(row=i, column=1).value == 31):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 32):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 33):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 34):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 35):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 36):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 37):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 38):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 39):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 40):
#         ws.cell(row=i, column=ripness).value = 0

#     elif(ws.cell(row=i, column=1).value == 41):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 42):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 43):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 44):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 45):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 46):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 47):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 48):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 49):
#         ws.cell(row=i, column=ripness).value = 0
#     elif(ws.cell(row=i, column=1).value == 50):
#         ws.cell(row=i, column=ripness).value = 0
    


wb.save(excel_filepath) # 若給予不同檔名代表另存新檔的意思