import cv2
import os
import numpy as np
import openpyxl
import numpy
from matplotlib import pyplot as plt
# import matplotlib.pyplot as plt
from tqdm import tqdm
from ellipse import ellipse
from otsu import yuv_otsu
from excel_clear import predict_excel_clear
from voting import voting
from PIL import Image
from openpyxl import load_workbook
import math
import sys   
import signal
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import pyshine as ps
import seaborn as sns
from sklearn import svm
from sklearn import datasets
from sklearn.datasets import load_iris
import joblib #https://ithelp.ithome.com.tw/articles/10197575
from sklearn.model_selection import train_test_split
from sklearn.decomposition import PCA
import warnings
warnings.filterwarnings("ignore")


# path='F:\dataset/19th_tomato/6th\sequoia\8l_test'
# predict_filepath='F:\dataset/19th_tomato\sequoia_mot/6th/tracks/img1.txt'
# k=path.split('/')
# r=predict_filepath.split('/')
# excel_path='F:\dataset/19th_tomato\sequoia_mot/6th/predict.xlsx' 
# video_path='F:\dataset/19th_tomato\sequoia_mot/6th/test.mp4'

dataset_name='19th2_tomato'
data=['9th','3l_test']

# path='F:\dataset/'+str(dataset_name)+'/'+str(data[0])+'\sequoia/'+str(data[1])
# predict_filepath='F:\dataset/'+str(dataset_name)+'\sequoia_mot/'+str(data[0])+'/tracks/img1.txt'
# k=path.split('/')
# r=predict_filepath.split('/')
# excel_path='F:\dataset/'+str(dataset_name)+'\sequoia_mot/'+str(data[0])+'/predict.xlsx' 
# video_path='F:\dataset/'+str(dataset_name)+'\sequoia_mot/'+str(data[0])+'/test.mp4'

path=os.path.join('F:\dataset',str(dataset_name),str(data[0]),'sequoia',str(data[1]))
predict_filepath=os.path.join('F:\dataset',str(dataset_name),'sequoia_mot',str(data[0]),'tracks','strongsort','img1.txt')
excel_path=os.path.join('F:\dataset',str(dataset_name),'sequoia_mot',str(data[0]),'predict.xlsx')
video_path=os.path.join('F:\dataset',str(dataset_name),'sequoia_mot',str(data[0]),'test.mp4')



# k=path.split('/')
# r=predict_filepath.split('/')
print(path)

# rgb_filepath=path+'/images/RGB'
rgb_filepath=path+'/images/RGB'
ndvi_filepath=path+'/NDVI'
ndre_filepath=path+'/NDRE'
gndvi_filepath=path+'/GNDVI'
grri_filepath=path+'/GRRI'

# k=path.split('/')

rgb_files=os.listdir(rgb_filepath)
ndvi_files=os.listdir(ndvi_filepath)
ndre_files=os.listdir(ndre_filepath)
gndvi_files=os.listdir(gndvi_filepath)
grri_files=os.listdir(grri_filepath)

#labels_filepath=rgb_filepath+'/labels_with_ids'
# excel_path='F:\warpping_try/try'+'.xls' 
# excel_path='F:\warpping_try/try'+'.xls'

# versus_component=['red','green','blue']
# versus_component=['ndvi','gndvi','grri']
versus_component=['red','green','blue','ndvi','gndvi','grri']
# versus_component=['red','green','r-g','g-b','hue','saturation','brightness','s-h','l','a','b','l-a']
#feature_columns = ['SepalLengthCm', 'SepalWidthCm', 'PetalLengthCm','PetalWidthCm']

classification_weight = joblib.load('F:\dataset/19th2_tomato\weights/svm/'+str(versus_component)+'.pkl')  # 選擇要使用的權重檔

workbook = openpyxl.Workbook()
#workbook = openpyxl.load_workbook(excel_path)
# 取得第一個工作表
sheet = workbook.worksheets[0]
# 設定 excel 工作表
excel_sheet=['frame','id','red','green','blue','r-g',
            'hue','saturation','brightness','s-h',
            'l','a','b','l-a',
            'ndvi','ndre','gndvi','grri','ripness',
            '','left','up','width','height']
sheet['A1'] = excel_sheet[0]
sheet['B1'] = excel_sheet[1]
sheet['C1'] = excel_sheet[2]
sheet['D1'] = excel_sheet[3]
sheet['E1'] = excel_sheet[4]
sheet['F1'] = excel_sheet[5]
sheet['G1'] = excel_sheet[6]
sheet['H1'] = excel_sheet[7]
sheet['I1'] = excel_sheet[8]
sheet['J1'] = excel_sheet[9]
sheet['K1'] = excel_sheet[10]
sheet['L1'] = excel_sheet[11]
sheet['M1'] = excel_sheet[12]
sheet['N1'] = excel_sheet[13]
sheet['O1'] = excel_sheet[14]
sheet['P1'] = excel_sheet[15]
sheet['Q1'] = excel_sheet[16]
sheet['R1'] = excel_sheet[17]
sheet['S1'] = excel_sheet[18]
sheet['T1'] = excel_sheet[19]
sheet['U1'] = excel_sheet[20]
sheet['V1'] = excel_sheet[21]
sheet['W1'] = excel_sheet[22]
sheet['X1'] = excel_sheet[23]
column=2
Correction_coefficient=0.9 #校正係數


# 儲存檔案
workbook.save(excel_path)
# def write_excel():

workbook.close()
tt=0
filtration_capacity=95
def filter(image,tt):
    img_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # 先轉成灰階處理
    maxium=np.percentile(img_gray, filtration_capacity, interpolation='midpoint') #https://www.796t.com/content/1550429124.html
    minimum=np.percentile(img_gray, (100-filtration_capacity)+100*(1-0.25*math.pi), interpolation='midpoint')
    #https://blog.csdn.net/JNingWei/article/details/77747959
    ret, thresh1 = cv2.threshold(img_gray, maxium, 255, 4)  # 利用 threshold 過濾出 #type = 4 ：此时小于阈值的像素点保持原色，大于阈值的像素点置0
    ret, thresh2 = cv2.threshold(img_gray, minimum, 255, 3)  # type = 3 ：此时大于阈值的像素点置填充色，小于阈值的像素点置0
    # concat_pic = np.concatenate([img_gray, thresh], axis=1)
    img = cv2.cvtColor(thresh1, cv2.COLOR_GRAY2BGR)  
    res = cv2.bitwise_and(image, image,mask=thresh1)
    res = cv2.bitwise_and(res, res,mask=thresh2)

    # if not os.path.isdir('./crop_images/lin/5th_95/'):
    #     os.mkdir('./crop_images/lin/5th_95/')
    # if(tt<=500):
    #     cv2.imwrite('./crop_images/lin/5th_95/'+str(tt)+'.jpg', res)
    
    return res

def mask(image,tt):
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    #部分2：
    # 在HSV空间中定义葉子部分
    lower_leaf = np.array([20, 0, 0])
    upper_leaf = np.array([180, 250, 200]) #要濾掉的
    # 在HSV空间中定义绿色
    lower_reflect = np.array([0, 0, 252])
    upper_reflect = np.array([180, 255, 255])  #要濾掉的
    # 在HSV空间中定义红色
    lower_red = np.array([0, 0, 160])
    upper_red = np.array([15, 255, 245])
    
    #部分3：
    # 从HSV图像中截取出蓝色、绿色、红色，即获得相应的掩膜
    # cv2.inRange()函数是设置阈值去除背景部分，得到想要的区域
    leaf = cv2.inRange(hsv, lower_leaf, upper_leaf)
    reflect = cv2.inRange(hsv, lower_reflect, upper_reflect)
    remained1 = cv2.bitwise_not(leaf)#逐位元not邏輯運算1
    remained2 = cv2.bitwise_not(reflect)#逐位元not邏輯運算1
    #部分4：
    # 将原图像和mask(掩膜)进行按位与
    remain=cv2.bitwise_and(frame,frame,mask=remained1)
    remain=cv2.bitwise_and(remain,remain,mask=remained2)
    
    #部分5:将BGR空间下的图片转换成RGB空间下的图片
    frame = frame[:,:,::-1]
    remain = remain[:,:,::-1]
    return remain




#切照片 
def crop_images(x,y,w,h,img):
    
    crop_img = img[y-int(0.5*h):y+int(0.5*h), x-int(0.5*w):x+int(0.5*w)]
    
    return crop_img
#切照片
def crop_tif_images(x,y,w,h,img):
    
    
    crop_img = img.crop((x-int(0.5*w),y-int(0.5*h),x+int(0.5*w),y+int(0.5*h)))
    return crop_img

pbar = tqdm(total=len(rgb_files),desc='計算各番茄的像素平均值')
for g in range (len(rgb_files)):
    # print(rgb_files[g][-4:])
    if(rgb_files[g][-3:]=='JPG' or rgb_files[g][-3:]=='jpg'):
        # try:
            f=open(predict_filepath,'r')
            rgb_image=cv2.imread(rgb_filepath+'/'+rgb_files[g])
            rgb_image=cv2.resize(rgb_image, (1280,960))
            # original_image=Image.open(rgb_filepath+'/'+rgb_files[g])
            # ndvi_image=Image.open(ndvi_filepath+'/'+ndvi_files[g])
            # ndre_image=Image.open(ndre_filepath+'/'+ndre_files[g])
            # gndvi_image=Image.open(gndvi_filepath+'/'+gndvi_files[g])
            # grri_image=Image.open(grri_filepath+'/'+grri_files[g])
            # nir_image=Image.open(nir_filepath+'/'+nir_files[g])
            
            ndvi_image=cv2.imread(ndvi_filepath+'/'+ndvi_files[g],2)
            ndre_image=cv2.imread(ndre_filepath+'/'+ndre_files[g],2)
            gndvi_image=cv2.imread(gndvi_filepath+'/'+gndvi_files[g],2)
            grri_image=cv2.imread(grri_filepath+'/'+grri_files[g],2)
           

              
            

            # original_size = original_image.size
            for line in f.readlines():
                s = line.split(',')
                s = list(map(int, s))#string轉int
                frame=s[0]
                id=s[1]
                width=int(s[4]*1280/4608)
                height=int(s[5]*960/3456)
                left=int(s[2]*1280/4608+(1-Correction_coefficient)/2*width)
                up=int(s[3]*960/3456+(1-Correction_coefficient)/2*height)
                width=int(s[4]*Correction_coefficient*(1280/4608))
                height=int(s[5]*Correction_coefficient*(960/3456))
                x_center=left+width/2
                y_center=up+height/2
                
                

                # id=int(s[1])                
                # x_center=float(s[2])*(ndvi_image.shape[1])
                # y_center=float(s[3])*(ndvi_image.shape[0])
                # width=float(s[4])*(ndvi_image.shape[1])*Correction_coefficient
                # height=float(s[5])*(ndvi_image.shape[0])*Correction_coefficient

                
                # image2=ellipse(image2)#切成橢圓
                # print(image2.shape)
                # print(x_center-0.5*width,y_center-0.5*height,x_center+0.5*width,y_center+0.5*height,ndvi_image.size[0],ndvi_image.size[1])
                

                # sndvi=ndre=0
                
                if((frame==g)&(left>=0)&(up>=0)&(0<(left+width)<rgb_image.shape[1])&(0<(up+height)<rgb_image.shape[0])):
                   

                    image2=crop_images(int(x_center),int(y_center),int(width),int(height),ndvi_image)
                    image3=crop_images(int(x_center),int(y_center),int(width),int(height),ndre_image)
                    image4=crop_images(int(x_center),int(y_center),int(width),int(height),gndvi_image)
                    image5=crop_images(int(x_center),int(y_center),int(width),int(height),grri_image)
                    # nir_image=crop_images(int(x_center),int(y_center),int(width),int(height),nir_image)
                    


                    # Correction_coefficient=0.7 #校正係數
                    
                    
                    # print(up,left,height,width)
                    image = rgb_image[up:up+height,left:left+width]
                    image=ellipse(image)#切成橢圓
                    # image=filter(image,tt)
                    image=yuv_otsu(image)

                    
                    tt=tt + 1
                    # print(id,x_center-0.5*width,y_center-0.5*height,x_center+0.5*width,y_center+0.5*height,original_image.shape[1],original_image.shape[0])
                    # size = image.shape
                    h, w = image.shape[:2]
                    # print(image.size)
                    # image=cv2.imread(rgb_filepath+'/'+rgb_files[g])
                    HSV=cv2.cvtColor(image,cv2.COLOR_BGR2HSV)
                    RGB=cv2.cvtColor(image,cv2.COLOR_BGR2RGB)
                    LAB=cv2.cvtColor(image,cv2.COLOR_BGR2LAB)
                    
                    
                    # cv2.imwrite('F:/dataset/19th_tomato/6th\sequoia/3l_train/crop/'+rgb_files[g][:-4]+'_'+s[1]+'.jpg',image)
                    # print('Size:'+str(size))
                    # print(size)
                    # exist = (image[0] != 0)
                    # area = exist.sum()
                    
                    area=0
                    red=green=blue=0
                    
                    hue=0
                    saturation=0
                    brightness=0
                    l=0
                    a=0
                    b=0
                    ndvi=0
                    ndre=0
                    gndvi=0
                    grri=0

                    
                     
                    for i in range (w):
                        for j in range(h):
                            
                                
                            if(RGB[j,i][0]!=0 and RGB[j,i][1]!=0 and RGB[j,i][2]!=0):
                                area=area+1
                                red=red+RGB[j,i][0]
                                green=green+RGB[j,i][1]
                                blue=blue+RGB[j,i][2]
                                # if(RGB[j,i][0]>0 and RGB[j,i][1]>0 and RGB[j,i][2]>0):
                                
                                #OpenCV uses H: 0-179, S: 0-255, V: 0-255
                                hue=hue+HSV[j,i][0]
                                saturation=saturation+HSV[j,i][1]
                                brightness=brightness+HSV[j,i][2]

                                l=l+LAB[j,i][0]
                                a=a+LAB[j,i][1]
                                b=b+LAB[j,i][2]

                                
                    # print(w,h)
                    # print(ndvi_image.size)
                    # image1 = numpy.array(ndvi_image)
                    

                    # image1=np.array(ndvi_image)
                    # image2=np.array(ndre_image)
                    # image3=np.array(gndvi_image)
                    # image4=np.array(grri_image)
                    
                    # image1 = cv2.cvtColor(ndvi_image, cv2.COLOR_BGR2GRAY)  
                    # image2 = cv2.cvtColor(ndre_image, cv2.COLOR_BGR2GRAY)
                    # image3 = cv2.cvtColor(gndvi_image, cv2.COLOR_BGR2GRAY)
                    # image4 = cv2.cvtColor(grri_image, cv2.COLOR_BGR2GRAY) 
                    # print('\n')
                    # print(h, w) 
                    # print(ndvi_image.shape)   
                    for i in range (w):
                        for j in range(h):
                            
                                
                            if(RGB[j,i][0]!=0 and RGB[j,i][1]!=0 and RGB[j,i][2]!=0):
  
                                ndvi=ndvi+image2[j,i]
                                ndre=ndre+image3[j,i]
                                gndvi=gndvi+image4[j,i]
                                grri=grri+image5[j,i]
                    # print('hello')

                    
                    # print(area)
                    if(area!=0):
                        averge_red=(red/area)
                        averge_green=(green/area)
                        averge_blue=(blue/area)
                        rgb_output=[averge_red,averge_green,averge_blue]
                        averge_hue=(hue/area)
                        averge_saturation=(saturation/area)
                        averge_brightness=(brightness/area)
                        hsv_output=[averge_hue,averge_saturation,averge_brightness]

                        averge_l=(l/area)
                        averge_a=(a/area)
                        averge_b=(b/area)

                        average_ndvi=ndvi/area*256
                        average_ndre=ndre/area*256
                        average_gndvi=gndvi/area*256
                        average_grri=grri/area*256
                        
                        # hsv_output=[averge_hue,averge_saturation,averge_brightness]

                        # wb = openpyxl.load_workbook(excel_path, data_only=True)
                        # data = [id,averge_red,averge_green,averge_blue,averge_hue,averge_saturation,averge_brightness]   # 二維陣列資料
                        # for i in data:
                        #     s3.append(i)

                        #將資料寫入excel
                        # print('q')
                        
                        sheet['A'+str(column)] =frame
                        sheet['B'+str(column)] =id
                        sheet['C'+str(column)] =averge_red
                        sheet['D'+str(column)] =averge_green
                        sheet['E'+str(column)] =averge_blue
                        sheet['F'+str(column)] =averge_red - averge_green
                        sheet['G'+str(column)] =averge_hue
                        sheet['H'+str(column)] =averge_saturation
                        sheet['I'+str(column)] =averge_brightness
                        sheet['J'+str(column)] =averge_saturation - averge_hue
                        sheet['K'+str(column)] =averge_l
                        sheet['L'+str(column)] =averge_a
                        sheet['M'+str(column)] =averge_b
                        sheet['N'+str(column)] =averge_l - averge_a
                        sheet['O'+str(column)] =average_ndvi
                        sheet['P'+str(column)] =average_ndre
                        sheet['Q'+str(column)] =average_gndvi
                        sheet['R'+str(column)] =average_grri
                                              
                        sheet['U'+str(column)] =s[2]
                        sheet['V'+str(column)] =s[3]
                        sheet['W'+str(column)] =s[4]
                        sheet['X'+str(column)] =s[5]
                        column=column+1
                    
                    
                    
                
               



                        
                    
        # except:
        #     pass


    workbook.save(excel_path)
    pbar.update()

predict_excel_clear(excel_path)

pbar.close()

# data_xls = pd.read_excel(excel_path)
# data_xls.to_csv('F:\dataset/19th_tomato/predict.csv', encoding='utf-8')
print('預測成熟度')
workbook = openpyxl.load_workbook(excel_path)
# test_data=pd.read_csv('F:\dataset/19th_tomato/predict.csv')
test_data = pd.read_excel(excel_path)
#https://ithelp.ithome.com.tw/articles/10197575


print('versus_component='+str(versus_component))

new = test_data[versus_component].values
predict_class = test_data['ripness'].values
pca2 = PCA(n_components=len(versus_component), iterated_power=1)
# pca = PCA(n_components=2, iterated_power=1)
# test_reduced = pca2.fit_transform(new)
predicted=classification_weight.predict(new)
sheet = workbook.worksheets[0]
column=2




for i in range(len(predicted)):
    sheet['S'+str(column)] =predicted[i]
    column=column+1
workbook.save(excel_path)

voting(excel_path)

print('生成影片')
filelist=os.listdir(rgb_filepath)
size = (4608,3456)
videowrite = cv2.VideoWriter(video_path,-1,10,size)
column=0
excel_data = pd.read_excel(excel_path)
predict_frame= excel_data['frame'].values
predict_id= excel_data['id'].values
predict_ripness = excel_data['ripness'].values
left = excel_data['left'].values
up = excel_data['up'].values
width = excel_data['width'].values
height = excel_data['height'].values

m=[0,0,0,0,0] #[0,-1,-2,-3,-4]

pbar = tqdm(total=len(filelist),desc='合成影片')
id_max=0
for i in range(len(filelist)):
    img=cv2.imread(rgb_filepath+'/'+filelist[i])  
    # id_background = np.zeros((150,300,3), dtype='uint8')
    for column in range (len(predict_frame)):
        if(predict_frame[column]==i):
            #cv2.rectangle(影像, 頂點座標, 對向頂點座標, 顏色, 線條寬度)
            #https://blog.gtwang.org/programming/opencv-drawing-functions-tutorial/
            #-1到-4分別為[紅、橙、黃、綠](b,g,r)
            if(predict_ripness[column]==-1):
                # cv2.putText(img, text, org, fontFace, fontScale, color, thickness, lineType)  #cv2.putText(img, text, org, fontFace, fontScale, color, thickness, lineType)
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 0, 255), 10)
                
            elif(predict_ripness[column]==-2):
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 165, 255), 10)
                
            elif(predict_ripness[column]==-3):
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 255, 255), 10)
                
            elif(predict_ripness[column]==-4):
                cv2.rectangle(img, (left[column], up[column]), (left[column]+width[column], up[column]+height[column]), (0, 255, 0), 10)
                
        if(predict_frame[column]==i):
            if(predict_id[column]>id_max):
                id_max=predict_id[column]
                # print(predict_id[column],predict_ripness[column])
                if(predict_ripness[column]==-1):
                    
                    m[1]=m[1]+1
                elif(predict_ripness[column]==-2):
                    
                    m[2]=m[2]+1
                elif(predict_ripness[column]==-3):
                    
                    m[3]=m[3]+1
                elif(predict_ripness[column]==-4):
                    
                    m[4]=m[4]+1
            
            # print(m)

            # 文字內容

    text = 'mature:'+str(m[1])+'  mature soon:'+str(m[2])+'  immature:'+str(m[3])
    text1 = 'mature:'+str(m[1])
    text2 = 'mature soon:'+str(m[2])
    text3 = 'immature:'+str(m[3])
    # # 顏色
    # color = (228, 225, 222)
    # text_color = (0, 255, 0)
    # x1=0
    # y1=0
    # # 計算文字的大小
    # (w, h), _ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 5.0, 1)

    # # 加入文字背景區塊
    # img = cv2.rectangle(img, (x1, y1 ), (x1 + w, y1 + h), color, -1)
    # # 加入文字
    # img = cv2.putText(img, text1, (x1, y1), cv2.FONT_HERSHEY_SIMPLEX, 5.0, text_color, 1)
        # text ='hello world'
    #https://officeguide.cc/python-opencv-add-background-box-with-text-tutorial-examples/
        # 加入文字方塊
    font_scale=10.0
    img = ps.putBText(
            img,                            # 原始影像
            text1,                             # 文字內容
            text_offset_x = 20,               # X 軸座標
            text_offset_y = 20,               # Y 軸座標
            vspace = 20,                      # 縱向空間
            hspace = 20,                      # 橫向空間
            font_scale = font_scale,                 # 字型大小
            background_RGB = (255, 255, 255), # 背景顏色
            text_RGB = (255,0,0),           # 文字顏色 (r,g,b)
            font = cv2.FONT_HERSHEY_PLAIN,    # 字型
            thickness = 15                    # 字體線條粗細
        )
    
    (w1, h1), _ = cv2.getTextSize(text1, cv2.FONT_HERSHEY_PLAIN, font_scale, 1)

    img = ps.putBText(
            img,                            # 原始影像
            text2,                             # 文字內容
            text_offset_x = 20 + w1 + 50,               # X 軸座標
            text_offset_y = 20,               # Y 軸座標
            vspace = 20,                      # 縱向空間
            hspace = 20,                      # 橫向空間
            font_scale =font_scale,                 # 字型大小
            background_RGB = (255, 255, 255), # 背景顏色
            text_RGB = (255, 165, 0),           # 文字顏色 
            font = cv2.FONT_HERSHEY_PLAIN,    # 字型
            thickness = 15                    # 字體線條粗細
        )
    (w2, h2), _ = cv2.getTextSize(text2, cv2.FONT_HERSHEY_PLAIN, font_scale, 1)
    img = ps.putBText(
            img,                            # 原始影像
            text3,                             # 文字內容
            text_offset_x = 20 + w1 + w2 + 150,               # X 軸座標
            text_offset_y = 20,               # Y 軸座標
            vspace = 20,                      # 縱向空間
            hspace = 20,                      # 橫向空間
            font_scale = font_scale,                 # 字型大小
            background_RGB = (255, 255, 255), # 背景顏色
            text_RGB = (255, 255, 0),           # 文字顏色 
            font = cv2.FONT_HERSHEY_PLAIN,    # 字型
            thickness = 15                    # 字體線條粗細
        )    

            
            
    videowrite.write(img)
    # m=[0,0,0,0,0]
    pbar.update()
pbar.close()
print('完成')